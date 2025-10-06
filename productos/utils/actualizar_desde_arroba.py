from openpyxl import load_workbook
from productos.models import Producto, ConfiguracionGlobal

def obtener_tasa_cambio():
    config = ConfiguracionGlobal.objects.first()
    return float(config.tasa_cambio_usd_mxn) if config and hasattr(config, 'tasa_cambio_usd_mxn') else 17.0

def actualizar_datos_desde_arroba(archivo_excel):
    tasa_cambio = obtener_tasa_cambio()
    print(f"✅ Tasa de cambio aplicada (Arroba): {tasa_cambio}")

    wb = load_workbook(archivo_excel, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        sku = str(row[0].value).strip() if row[0].value else None
        precio_mxn = row[6].value
        moneda = row[7].value
        stock = row[10].value

        if not sku:
            continue

        try:
            producto = Producto.objects.get(sku=sku)

            if moneda == "USD":
                precio_final = float(precio_mxn) * tasa_cambio
            else:
                precio_final = float(precio_mxn)

            producto.precio_arroba = round(precio_final, 2)
            producto.stock_arroba = int(stock) if stock is not None else 0
            producto.save()
            print(f"✔️ Producto actualizado (Arroba): {sku}")
        except Producto.DoesNotExist:
            print(f"❌ SKU no encontrado (Arroba): {sku}")

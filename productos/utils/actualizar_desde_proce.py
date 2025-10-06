from openpyxl import load_workbook
from productos.models import Producto, ConfiguracionGlobal

def obtener_tasa_cambio():
    config = ConfiguracionGlobal.objects.first()
    return float(config.tasa_cambio_usd_mxn) if config and hasattr(config, 'tasa_cambio_usd_mxn') else 17.0

def actualizar_datos_desde_proce(archivo_excel):
    print("✅ Procesando catálogo Proce...")
    wb = load_workbook(archivo_excel, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        sku = str(row[0].value).strip() if row[0].value else None
        precio = row[5].value
        stock = row[4].value

        if not sku:
            continue

        try:
            producto = Producto.objects.get(sku=sku)
            producto.precio_proce = float(precio) if precio else 0
            producto.stock_proce = int(stock) if stock else 0
            producto.save()
            print(f"✔️ Producto actualizado (Proce): {sku}")
        except Producto.DoesNotExist:
            print(f"❌ SKU no encontrado (Proce): {sku}")
